import time
import asyncio
from playwright.async_api import async_playwright
from playwright._impl._errors import TimeoutError, Error
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

async def measure_page_load_time(url: str, refresh_times: int):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()
        # Pre-launch the browser instance
        print('Initializing the browser...')
        await page.goto(url, wait_until='networkidle') 
        # Warm-up page load: 
        # The first time the browser is launched
        # it usually takes longer due to the time required for initializing the browser instance
        # This warm-up load is performed to ensure the browser is fully initialized
        # but we won't measure or record its load time
        # The actual page load times will be measured in subsequent refreshes.

        load_times = []
        for i in range(refresh_times):
            await context.clear_cookies()
            print('Page Loading...')
            start_time = time.time()

            try:
                await page.goto(url, wait_until='networkidle', timeout=30000)
            except (asyncio.TimeoutError, TimeoutError, Error) as e:
                if "net::ERR_CONNECTION_REFUSED" in str(e):
                     print("Check if the URL is valid.")
                     exit(1)
                else:
                    print(f"Page load time for refresh {i + 1} exceeded 30 seconds. Skipping...")
                    # print(e)
                    load_times.append('OT')
                    continue
            end_time = time.time()

            load_time = end_time - start_time

            print(f"Page load time for refresh {i + 1}: {load_time:.2f} seconds")
            load_times.append(load_time)

        # Save the results to an Excel file
        save_info_to_excel(url, refresh_times, load_times)
        # Close the browser
        await browser.close()

def save_info_to_excel(url: str, refresh_times: int, load_times: list):
    try:
        wb = load_workbook('page_load_time.xlsx')
    except FileNotFoundError:
        wb = Workbook()
    
    timestamp = datetime.datetime.now().strftime('%d-%m-%Y-%H%M%S%f')[:-3]
    
    sheet_name = f'{timestamp}'
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
    
    sheet.append(['Refresh Times', 'URL', 'Load Times'])
    for i in range(refresh_times):
        sheet.append([i + 1, url, load_times[i]])
    # Calculate the average load time
    non_ot_load_times = [time for time in load_times if time != 'OT']
    if non_ot_load_times:
        average_load_time = sum(non_ot_load_times) / len(non_ot_load_times)
        print(f"Average page load time: {average_load_time:.2f} seconds")
    else:
        print("No successful page loads.")
    sheet.append(['', 'Average page load time: ', average_load_time])
    wb.save('page_load_time.xlsx')

# Get user input for URL and refresh times
if __name__ == "__main__":
    url = input("Enter the URL of the page to measure: ")
    refresh_times = int(input("Enter the number of times to refresh the page: "))
    asyncio.run(measure_page_load_time(url, refresh_times))
